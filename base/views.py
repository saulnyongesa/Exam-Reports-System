import base64
from datetime import timezone
from django.utils.timezone import now, localtime

import datetime
from io import BytesIO
import json
import random
from django.views.decorators.csrf import csrf_exempt
from django.contrib.auth.decorators import login_required
from django.contrib.auth import authenticate, login, logout
from django.shortcuts import render
from django.contrib import messages
from openpyxl import load_workbook
from reports.forms import *
from reports.models import *
from django.shortcuts import get_object_or_404, redirect
from django.urls import reverse
import csv
from django.db.models import Q
from django.http import HttpResponse, HttpResponseRedirect, JsonResponse
from django.core.files.storage import default_storage
import numpy as np
import cv2
import face_recognition
from django.core.mail import send_mail
from django.contrib.auth.hashers import make_password

# USER ACCOUNTS=============================================================================================


def generate_otp():
    return random.randint(100000, 999999)


def signin(request):
    if request.user.is_authenticated:
        return redirect("teacher-dashboard-url")
    else:
        if request.method == "POST":
            email = request.POST.get("email")
            password = request.POST.get("password")
            user = authenticate(request, username=email, password=password)
            if user is not None and user.is_staff:
                otp = OTP.objects.create(user=user, otp=generate_otp())
                otp.save()
                send_mail(
                    "Med Lear Account Login Verification",
                    f"Your verification code is: {otp.otp} Expires in 5 minutes.",
                    "sanymtechs@gmail.com",
                    [email],
                    fail_silently=False,
                )
                messages.success(
                    request, "Check your email for OTP. It Expires in 5 minutes."
                )
                return HttpResponseRedirect("/account/verify/" + str(user.id) + "/")
            else:
                messages.info(request, "Email or password is incorrect")
    return render(request, "pwd/sign_in.html")


def otp_verify(request, id):
    user = User.objects.get(id=id)
    if request.method == "POST":
        otp_code = request.POST.get("otp_code")
        try:
            otp = OTP.objects.get(user=user, otp=otp_code)
            if otp:
                if request.user.is_authenticated:
                    return redirect("report-home-url")
                if user is not None and user.is_staff:
                    login(request, user)
                    otp.delete()
                    messages.success(request, "Welcome")
                    return redirect("report-home-url")

        except OTP.DoesNotExist:
            messages.error(request, "Inavalid otp")
            return HttpResponseRedirect("/account/verify/" + str(user.id) + "/")
    context = {"user": user}
    return render(request, "pwd/otp.html", context)


@login_required(login_url="signin-url")
def sign_out_user(request):
    logout(request)
    return redirect("signin-url")


# ADMIN PAGE=============================================================================================
def home(request):
    return render(request, "base/home.html")


@login_required(login_url="signin-url")
def profile(request):
    form = UserForm(instance=request.user)
    return render(request, "base/profile.html", {"form_profile": form})


@login_required(login_url="signin-url")
def profile_edit(request):
    if request.method == "POST":
        form = UserForm(request.POST, instance=request.user)
        if form.is_valid():
            form.save()
            messages.success(request, "You edited your profile successfully")
            return redirect("admin-profile-url")
        else:
            messages.error(request, "Please correct the errors")
    else:
        form = UserForm(instance=request.user)
    return render(request, "base/profile.html", {"form_profile": form})


# Other Admin Pages=============================================================================================
@login_required(login_url="clock-face-url")
def admin_signup(request):
    if request.method == "POST":
        form = AdminForm(request.POST)
        if form.is_valid():
            form.instance.is_staff = True
            form.instance.password = make_password(form.instance.id_number)
            form.instance.username = form.instance.email
            form.save()
            send_mail(
                "Reports MIS Account Creation",
                f"Your account has been created.Your username is: {form.instance.email} and Your password is your ID number.",
                "sanymtechs@gmail.com",
                [form.instance.email],
                fail_silently=False,
            )
            messages.success(request, "Admin acccount added successful")
            return redirect("admin-home-url")
    else:
        form = AdminForm()
    context = {
        "form": form,
    }
    return render(request, "base/admin_signup.html", context)


def admins_all(request):
    admins = User.objects.filter(is_staff=True)
    return render(request, "base/admins.html", {"admins": admins})


def admin_status(request, admin_id):
    admin = get_object_or_404(User, id=admin_id)
    if request.method == "POST":
        admin.is_active = not admin.is_active
        admin.save()
        messages.success(request, "Status changed successfully")
        return JsonResponse({"success": True, "message": "Status changed successfully"})
    return JsonResponse({"success": False, "message": "Failed to change status"})


def admin_delete(request, admin_id):
    admin = get_object_or_404(User, id=admin_id)
    if request.method == "DELETE":
        admin.delete()
        messages.success(request, "Admin deleted successfully")
        return JsonResponse({"success": True, "message": "Admin deleted successfully"})
    return JsonResponse({"success": False, "message": "Failed to delete admin"})


# UNIT AND COURSE=============================================================================================
def courses_all(request):
    courses = Course.objects.all()
    return render(request, "base/course_units/courses.html", {"courses": courses})


def course_detail(request, course_id):
    course = get_object_or_404(Course, id=course_id)
    return render(request, "base/course_units/course_details.html", {"course": course})


def add_course(request):
    if request.method == "POST":
        form = CourseForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect("admin-courses-url")
    else:
        form = CourseForm()
    return render(request, "base/course_units/course_add.html", {"form": form})


def edit_course(request, course_id):
    form = CourseForm(request.POST)
    course = get_object_or_404(Course, id=course_id)
    if request.method == "POST":
        form = CourseForm(request.POST, instance=course)
        if form.is_valid():
            form.save()
            return redirect("report-courses-url")
    else:
        form = CourseForm(instance=course)
    return render(request, "base/course_units/course_edit.html", {"form": form})


def delete_course(request, course_id):
    course = get_object_or_404(Course, id=course_id)
    if request.method == "DELETE":
        course.delete()
        return JsonResponse({"success": True, "message": "Course deleted successfully"})
    return JsonResponse({"success": False, "message": "Failed to delete course"})


def units_all(request):
    units = Unit.objects.all()
    return render(request, "base/course_units/units.html", {"units": units})


def unit_detail(request, unit_id):
    unit = get_object_or_404(Unit, id=unit_id)
    return render(request, "base/course_units/unit_details.html", {"unit": unit})


def add_unit(request):
    if request.method == "POST":
        form = UnitForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect("report-units-url")
    else:
        form = UnitForm()
    return render(request, "base/course_units/unit_add.html", {"form": form})


def edit_unit(request, unit_id):
    unit = get_object_or_404(Unit, id=unit_id)
    if request.method == "POST":
        form = UnitForm(request.POST, instance=unit)
        if form.is_valid():
            form.save()
            return redirect("report-units-url")
    else:
        form = UnitForm(instance=unit)
    return render(request, "base/course_units/unit_edit.html", {"form": form})


def delete_unit(request, unit_id):
    unit = get_object_or_404(Unit, id=unit_id)
    if request.method == "DELETE":
        unit.delete()
        return JsonResponse({"success": True, "message": "Unit deleted successfully"})
    return JsonResponse({"success": False, "message": "Failed to delete unit"})


def course_units_view(request):
    courses = Course.objects.all
    course_units = CourseUnit.objects.all()
    units = Unit.objects.all()
    context = {"courses": courses, "course_units": course_units, "units": units}
    return render(request, "base/course_units/courses_units.html", context)


def get_unlinked_units(request, course_id):
    linked_units = CourseUnit.objects.filter(course_id=course_id).values_list(
        "unit_id", flat=True
    )
    unlinked_units = Unit.objects.exclude(id__in=linked_units)
    return JsonResponse(
        {"units": list(unlinked_units.values("id", "unit_name", "unit_code"))}
    )


@csrf_exempt
def save_course_units(request):
    if request.method == "POST":
        try:
            data = json.loads(request.body)
            course_id = data.get("course_id")
            unit_ids = data.get("units", [])

            # Get the course instance
            course = get_object_or_404(Course, id=course_id)

            # Link the selected units to the course
            for unit_id in unit_ids:
                unit = get_object_or_404(Unit, id=unit_id)
                # Create a CourseUnit relationship
                CourseUnit.objects.get_or_create(course=course, unit=unit)

            return JsonResponse(
                {"success": True, "message": "Units saved successfully."}
            )
        except Exception as e:
            return JsonResponse({"success": False, "error": str(e)}, status=400)
    return JsonResponse(
        {"success": False, "message": "Invalid request method."}, status=405
    )


def delete_course_unit(request, id):
    course_unit = get_object_or_404(CourseUnit, id=id)
    if request.method == "DELETE":
        course_unit.delete()
        return JsonResponse({"success": True, "message": "Unit deleted successfully"})
    return JsonResponse({"success": False, "message": "Failed to delete unit"})


# IMPORT UNITS, COURSES===================================================================================
def import_units_from_excel(request):
    if not request.user.is_authenticated:
        messages.error(request, "You are not authenticated to perform this action")
        return redirect("admin-home-url")

    if request.method == "POST" and request.FILES.get("file"):
        excel_file = request.FILES["file"]
        wb = load_workbook(filename=BytesIO(excel_file.read()))
        ws = wb.active

        for row in ws.iter_rows(min_row=2, values_only=True):  # Skip row 1 (titles)
            row_values = list(row)
            if len(row_values) < 5 or not all(row_values):
                continue  # Skip rows with missing values

            (
                unit_code,
                unit_name,
                teaching_hrs_per_class,
                teaching_hrs_week,
                teaching_hrs_term,
            ) = row_values[:5]

            if not Unit.objects.filter(unit_code=unit_code).exists():
                Unit.objects.create(
                    unit_name=unit_name,
                    unit_code=unit_code,
                    teaching_hrs_per_class=teaching_hrs_per_class,
                    teaching_hrs_week=teaching_hrs_week,
                    teaching_hrs_term=teaching_hrs_term,
                )

        messages.success(request, "Units imported successfully")
        return redirect("admin-units-url")

    return render(request, "base/course_units/import_units.html")


def import_courses_from_excel(request):
    if not request.user.is_authenticated:
        messages.error(request, "You are not authenticated to perform this action")
        return redirect("admin-home-url")

    if request.method == "POST" and request.FILES.get("file"):
        excel_file = request.FILES["file"]
        wb = load_workbook(filename=BytesIO(excel_file.read()))
        ws = wb.active

        for row in ws.iter_rows(min_row=2, values_only=True):  # Skip row 1 (titles)
            row_values = list(row)
            if len(row_values) < 2 or not all(row_values):
                continue  # Skip rows with missing values

            course_code, course_name = row_values[:2]

            if not Course.objects.filter(course_code=course_code).exists():
                Course.objects.create(course_name=course_name, course_code=course_code)

        messages.success(request, "Courses imported successfully")
        return redirect("admin-courses-url")

    return render(request, "base/course_units/import_courses.html")


def trainer_signup(request):
    form = TrainerSignupForm()

    if request.method == "POST":
        form = TrainerSignupForm(request.POST, request.FILES)

        if form.is_valid():
            trainer = form.save(commit=False)
            uploaded_photo = request.FILES.get("photo")

            if uploaded_photo:
                # Save the uploaded photo
                file_name = default_storage.save(uploaded_photo.name, uploaded_photo)
                file_path = default_storage.path(file_name)

                # Load image using OpenCV
                frame = cv2.imread(file_path)

                if frame is None:
                    messages.error(
                        request,
                        "Could not read the image file. Ensure it is a valid image.",
                    )
                    return render(
                        request, "trainer/trainer_signup.html", {"form": form}
                    )

                # Convert to RGB format
                frame = cv2.cvtColor(frame, cv2.COLOR_BGR2RGB)
                # Check the image type
                if frame.dtype != np.uint8:
                    frame = frame.astype(np.uint8)

                # Detect face encodings
                face_encodings = face_recognition.face_encodings(frame)

                if face_encodings:
                    face_encoding = face_encodings[0]
                    trainer.face_encoding = ",".join(map(str, face_encoding))
                    trainer.save()
                    messages.success(request, "Trainer account created successfully")
                    return redirect("trainers-url")
                else:
                    messages.error(
                        request,
                        "No face detected in the uploaded photo. Make sure the photo is clear and contains only one face.",
                    )
            else:
                messages.error(request, "Please upload a photo. It is required.")

    return render(request, "trainer/trainer_signup.html", {"form": form})


def trainers_all(request):
    trainers = Trainer.objects.all()
    return render(request, "trainer/trainers.html", {"trainers": trainers})


def trainer_status(request, trainer_id):
    trainer = get_object_or_404(Trainer, id=trainer_id)
    if request.method == "POST":
        trainer.is_active = not trainer.is_active
        trainer.save()
        messages.success(request, "Status changed successfully")
        return JsonResponse({"success": True, "message": "Status changed successfully"})
    return JsonResponse({"success": False, "message": "Failed to change status"})


def trainer_delete(request, trainer_id):
    trainer = get_object_or_404(Trainer, id=trainer_id)
    if request.method == "DELETE":
        trainer.delete()
        messages.success(request, "Trainer deleted successfully")
        return JsonResponse({"success": True, "message": "Admin deleted successfully"})
    return JsonResponse({"success": False, "message": "Failed to delete admin"})


def trainer_details(request, trainer_id):
    trainer = get_object_or_404(Trainer, id=trainer_id)
    trainer_units = TrainerUnit.objects.filter(trainer=trainer)
    assigned_unit_ids = trainer_units.values_list("unit_id", flat=True)
    unassigned_units = Unit.objects.exclude(id__in=assigned_unit_ids)
    return render(
        request,
        "trainer/trainer.html",
        {"trainer": trainer, "trainer_units": trainer_units, "units": unassigned_units},
    )


def trainer_edit(request, trainer_id):
    trainer = get_object_or_404(Trainer, id=trainer_id)
    if request.method == "POST":
        form = TrainerSignupForm(request.POST, request.FILES, instance=trainer)
        if form.is_valid():
            trainer = form.save(commit=False)
            uploaded_photo = request.FILES.get("photo")

            if uploaded_photo:
                # Save the uploaded photo
                file_name = default_storage.save(uploaded_photo.name, uploaded_photo)
                file_path = default_storage.path(file_name)

                # Load image using OpenCV
                frame = cv2.imread(file_path)

                if frame is None:
                    messages.error(
                        request,
                        "Could not read the image file. Ensure it is a valid image.",
                    )
                    return render(
                        request, "trainer/trainer_signup.html", {"form": form}
                    )

                # Convert to RGB format
                frame = cv2.cvtColor(frame, cv2.COLOR_BGR2RGB)
                # Check the image type
                if frame.dtype != np.uint8:
                    frame = frame.astype(np.uint8)

                # Detect face encodings
                face_encodings = face_recognition.face_encodings(frame)

                if face_encodings:
                    face_encoding = face_encodings[0]
                    trainer.face_encoding = ",".join(map(str, face_encoding))
                    trainer.save()
                    messages.success(request, "Trainer Details Edited successfully")
                    return HttpResponseRedirect(
                        reverse("trainer-url", args=[trainer_id])
                    )
                else:
                    messages.error(
                        request,
                        "No face detected in the uploaded photo. Make sure the photo is clear and contains only one face.",
                    )
            else:
                messages.error(request, "Please upload a photo. It is required.")
    else:
        form = TrainerSignupForm(instance=trainer)
    return render(
        request, "trainer/trainer_edit.html", {"form": form, "trainer": trainer}
    )


def get_unlinked_units_to_trainer(request, trainer_id):
    trainer_units = TrainerUnit.objects.filter(trainer_id=trainer_id).values_list(
        "unit_id", flat=True
    )
    unlinked_units = Unit.objects.exclude(id__in=trainer_units)
    data = {
        "units": [
            {"id": unit.id, "unit_name": unit.unit_name, "unit_code": unit.unit_code}
            for unit in unlinked_units
        ]
    }
    return JsonResponse(data)


def save_trainer_units(request):
    if request.method == "POST":
        import json

        data = json.loads(request.body)
        trainer_id = data.get("trainer_id")
        unit_ids = data.get("units", [])
        trainer = get_object_or_404(Trainer, id=trainer_id)
        for unit_id in unit_ids:
            unit = get_object_or_404(Unit, id=unit_id)
            TrainerUnit.objects.get_or_create(trainer=trainer, unit=unit)
        return JsonResponse({"success": True})
    return JsonResponse({"success": False, "error": "Invalid request"}, status=400)


def trainer_unit_delete(request, id):
    trainer = get_object_or_404(TrainerUnit, id=id)
    if request.method == "DELETE":
        trainer.delete()
        messages.success(request, "Unit removed from the trainer successfully")
        return JsonResponse(
            {"success": True, "message": "Unit removed from the trainer successfully"}
        )
    return JsonResponse(
        {
            "success": False,
            "message": "Failed to remove Unit from the trainer successfully",
        }
    )


# Clock in and out admin side=============================================================================================
def clock(request):
    if request.method == "POST":
        id_number = request.POST.get("id_number")
        try:
            trainer = Trainer.objects.get(id_number=id_number)
            if trainer and trainer.is_active:
                try:
                    teaching_attendance = TeachingAttendance.objects.get(
                        trainer=trainer, is_clocked_in=True, is_clocked_out=False
                    )
                    messages.error(
                        request,
                        f"Trainer is already clocked in for {teaching_attendance.unit.unit_name} ({teaching_attendance.unit.unit_code})",
                    )
                    return redirect("admin-clock-out-url", trainer_id=trainer.id)
                except TeachingAttendance.DoesNotExist:
                    return redirect("admin-clock-in-url", trainer_id=trainer.id)
        except Trainer.DoesNotExist:
            messages.error(request, "Trainer not found or the Trainer is not active")
            return redirect("admin-clock-url")
    return render(request, "base/clock/clock.html")
# Clock in and out trainer side=============================================================================================

def clock_out(request, trainer_id):
    trainer = get_object_or_404(Trainer, id=trainer_id)
    teaching_attendance = TeachingAttendance.objects.get(
        trainer=trainer, is_clocked_in=True, is_clocked_out=False
    )
    (
        hours_covered,
        percentage_hours_week,
        expected_hours_week,
        remaining_hours,
        expected_hours_per_class,
    ) = calculate_teaching_hours(trainer, teaching_attendance.unit)
    if request.method == "POST":
        roll = request.POST.get("roll")
        teaching_attendance.roll = roll
        teaching_attendance.clock_out = now()
        teaching_attendance.is_clocked_out = True
        teaching_attendance.save()
        messages.success(request, "Trainer clocked out successfully")
        return redirect("admin-clock-url")
    context = {
        "trainer": trainer,
        "teaching_attendance": teaching_attendance,
        "expected_hours_per_week": expected_hours_week,
        "expected_hours_per_class": expected_hours_per_class,
        "hours_covered": hours_covered,
        "percentage": percentage_hours_week,
        "remaining_hours": remaining_hours,
    }
    return render(request, "base/clock/clock_out.html", context)


def clock_in(request, trainer_id):
    trainer = get_object_or_404(Trainer, id=trainer_id)
    units = Unit.objects.filter(trainerunit__trainer=trainer)
    if request.method == "POST":
        unit_id = request.POST.get("unit_id")
        unit = get_object_or_404(Unit, id=unit_id)
        teaching_attendance = TeachingAttendance.objects.create(
            trainer=trainer, unit=unit, clock_in=now()
        )
        teaching_attendance.save()
        messages.success(
            request,
            f"Trainer clocked in for {unit.unit_name} ({unit.unit_code}) successfully",
        )
        return redirect("admin-clock-url")
    context = {"trainer": trainer, "units": units}
    return render(request, "base/clock/clock_in.html", context)



# Clock in and out trainer side=============================================================================================

def trainer_clock(request):
    return render(request, "base/clock/trainer_clock.html")

def face_validation(request):
    if request.method == 'POST':
        data = json.loads(request.body)
        image_data = data.get('image_data')
        image_bytes = base64.b64decode(image_data.split(',')[1])
        nparr = np.frombuffer(image_bytes, np.uint8)
        frame = cv2.imdecode(nparr, cv2.IMREAD_COLOR)
        face_encodings = face_recognition.face_encodings(frame)

        if face_encodings:
            best_match = None
            lowest_distance = float('inf')
            for trainer in Trainer.objects.all():
                try:
                    known_encoding = np.fromstring(trainer.face_encoding, dtype=float, sep=',')
                    faceDistances = face_recognition.face_distance([known_encoding], face_encodings[0])
                    if faceDistances[0] < lowest_distance:
                        lowest_distance = faceDistances[0]
                        best_match = trainer
                except Exception as e:
                    messages.error(request, 'Face not detected or trainer not registered')
            if best_match and lowest_distance <= 0.6:
                return JsonResponse({'success': True, 'trainer_name': best_match.name, 'id': best_match.id}, )
            else:
                return JsonResponse({'success': False, 'message': 'Face not recognized'})
        else:
            return JsonResponse({'success': False, 'message': 'No face detected'})
    else:
        return JsonResponse({'success': False, 'message': 'Invalid request'})

def trainer_clock_afater_validtion(request, trainer_id):
        try:
            trainer = Trainer.objects.get(id=trainer_id)
            if trainer and trainer.is_active:
                try:
                    teaching_attendance = TeachingAttendance.objects.get(
                        trainer=trainer, is_clocked_in=True, is_clocked_out=False
                    )
                    messages.error(
                        request,
                        f"You already clocked in for {teaching_attendance.unit.unit_name} ({teaching_attendance.unit.unit_code})",
                    )
                    return redirect("trainer-clock-out-url", trainer_id=trainer.id)
                except TeachingAttendance.DoesNotExist:
                    return redirect("trainer-clock-in-url", trainer_id=trainer.id)
        except Trainer.DoesNotExist:
            messages.error(request, "Trainer not found or the Trainer is not active")
            return redirect("trainer-clock-url")

def trainer_clock_out(request, trainer_id):
    trainer = get_object_or_404(Trainer, id=trainer_id)
    teaching_attendance = TeachingAttendance.objects.get(
        trainer=trainer, is_clocked_in=True, is_clocked_out=False
    )
    (
        hours_covered,
        percentage_hours_week,
        expected_hours_week,
        remaining_hours,
        expected_hours_per_class,
    ) = calculate_teaching_hours(trainer, teaching_attendance.unit)
    if request.method == "POST":
        roll = request.POST.get("roll")
        teaching_attendance.roll = roll
        teaching_attendance.clock_out = now()
        teaching_attendance.is_clocked_out = True
        teaching_attendance.save()
        messages.success(request, "Trainer clocked out successfully")
        return redirect("trainer-clock-url")
    context = {
        "trainer": trainer,
        "teaching_attendance": teaching_attendance,
        "expected_hours_per_week": expected_hours_week,
        "expected_hours_per_class": expected_hours_per_class,
        "hours_covered": hours_covered,
        "percentage": percentage_hours_week,
        "remaining_hours": remaining_hours,
    }
    return render(request, "base/clock/trainer_clock_out.html", context)


def trainer_clock_in(request, trainer_id):
    trainer = get_object_or_404(Trainer, id=trainer_id)
    units = Unit.objects.filter(trainerunit__trainer=trainer)
    if request.method == "POST":
        unit_id = request.POST.get("unit_id")
        id_number = request.POST.get("id_number")
        if id_number != trainer.id_number:
            messages.error(request, "ID Number did not match! Try again")
            return redirect("trainer-clock-in-url", trainer_id=trainer.id)
        else:
            unit = get_object_or_404(Unit, id=unit_id)
            teaching_attendance = TeachingAttendance.objects.create(
                trainer=trainer, unit=unit, clock_in=now()
            )
            teaching_attendance.save()
            messages.success(
                request,
                f"You have clocked in for {unit.unit_name} ({unit.unit_code}) successfully",
            )
            return redirect("trainer-clock-url")
    context = {"trainer": trainer, "units": units}
    return render(request, "base/clock/trainer_clock_in.html", context)


def calculate_teaching_hours(trainer, unit):
    current_time = localtime(now())  # Ensures local timezone
    start_of_week = current_time - timedelta(days=current_time.weekday())  # Get Monday
    start_of_week = start_of_week.replace(hour=0, minute=0, second=0, microsecond=0)  # Midnight
    if is_naive(start_of_week):
        start_of_week = make_aware(start_of_week)
    try:
        trainer_unit = TrainerUnit.objects.get(trainer=trainer, unit=unit)
    except TrainerUnit.DoesNotExist:
        return None, None, None, None
    teaching_attendances = TeachingAttendance.objects.filter(
        trainer=trainer_unit.trainer,
        unit=trainer_unit.unit,
        clock_in__gte=start_of_week,
        clock_in__lte=current_time,
    )
    expected_hours = trainer_unit.unit.teaching_hrs_week or 1
    expected_hours_per_class_ = trainer_unit.unit.teaching_hrs_per_class or 1
    total_minutes = (
        sum(
            attendance.time_taken
            for attendance in teaching_attendances
            if attendance.time_taken
        )
        or 0
    )
    total_hours = total_minutes // 60
    rem_minutes = total_minutes % 60

    percentage_hours_covered = (
        (total_hours * 100) / expected_hours if expected_hours else 0
    )
    remaining_hours = expected_hours - total_hours

    return (
        f"{total_hours}Hrs {rem_minutes} Minutes",
        percentage_hours_covered,
        expected_hours,
        remaining_hours,
        expected_hours_per_class_,
    )


def get_unit_data(request):
    if request.method == "POST":
        try:
            data = json.loads(request.body)
            unit_id = data.get("unit_id")
            trainer_id = data.get("trainer_id")

            if not unit_id or not trainer_id:
                return JsonResponse(
                    {"success": False, "message": "Invalid Values Given"}
                )

            trainer = Trainer.objects.get(id=trainer_id)

            try:
                unit = Unit.objects.get(id=unit_id)
            except unit.DoesNotExist:
                return JsonResponse({"success": False, "message": "Unit Not Found"})

            (
                hours_covered,
                percentage_hours_week,
                expected_hours_week,
                remaining_hours,
                expected_hours_per_class,
            ) = calculate_teaching_hours(trainer, unit)

            if hours_covered is None:
                return JsonResponse(
                    {"success": False, "message": "No attendance data found"}
                )

            context = {
                "success": True,
                "unit_name": unit.unit_name,
                "unit_code": unit.unit_code,
                "expected_hours": expected_hours_week,
                "expected_hours_per_class": expected_hours_per_class,
                "hours_covered": hours_covered,
                "percentage": percentage_hours_week,
                "remaining_hours": remaining_hours,
            }
            return JsonResponse(context)

        except json.JSONDecodeError:
            return JsonResponse({"success": False, "message": "Invalid JSON format"})

    return JsonResponse({"success": False, "message": "Invalid request"})


# Clock in/ou History and reports======================================================================================
def clock_reports(request):
    return render(request, 'base/clock/clock_reports.html')

def clock_history_trainer(request):
        if request.method == 'POST':
            id_number = request.POST.get('id-number').upper()
            start_date = request.POST.get('date1')
            end_date = request.POST.get('date2')
            start_date = datetime.strptime(start_date, '%Y-%m-%d')
            end_date = datetime.strptime(end_date, '%Y-%m-%d')
            DatePicker.start_date = start_date
            DatePicker.end_date=end_date
            print(DatePicker.start_date)
            try:
                trainer = Trainer.objects.get(id_number=id_number)
                if trainer:
                    clocks = TeachingAttendance.objects.filter(
                        Q(clock_in__date__gte=start_date) &
                        Q(clock_in__date__lte=end_date) &
                        Q(trainer=trainer)
                    )
                    if clocks:
                        context = {
                            'clocks': clocks,
                            'trainer': trainer,
                            'start': start_date,
                            'end': end_date
                        }
                        return render(request, 'base/clock/clock_history.html', context)
                    else:
                        messages.error(request, 'No information found')
                        return redirect('clock-history-trainer-url')
            except Trainer.DoesNotExist:
                messages.error(request, f'Trainer with ID Number: {id_number} not found')
                return redirect('clock-history-trainer-url')
        return render(request, 'base/clock/clock_history.html')

def clock_history_all(request):
        if request.method == 'POST':
            start_date = request.POST.get('date1')
            end_date = request.POST.get('date2')
            try:
                    start_date = datetime.strptime(start_date, '%Y-%m-%d')
                    end_date = datetime.strptime(end_date, '%Y-%m-%d')
                    DatePicker.start_date = start_date
                    DatePicker.end_date=end_date
                    clocks = TeachingAttendance.objects.filter(
                        Q(clock_in__date__gte=start_date) & Q(clock_in__date__lte=end_date)
                    )
                    context = {
                        "clocks": clocks,
                        'start': start_date,
                        'end': end_date
                    }
                    return render(request, 'base/clock/clock_history_all.html', context)
            except TeachingAttendance.DoesNotExist:
                    messages.error(request, 'No information found')
                    return redirect('clock-history-all-url')
        return render(request, 'base/clock/clock_history_all.html')


def active_classes(request):
        clocks = TeachingAttendance.objects.filter(
            Q(is_clocked_out=False)
        )
        return render(request, 'base/clock/clock_active_classes.html', {'clocks': clocks})
    
class DatePicker:
    start_date = ''
    end_date = ''
    
