from io import BytesIO
from django.http import HttpResponse
from django.shortcuts import redirect, render
from django.contrib import messages
from openpyxl import Workbook
from openpyxl.styles import Font
from reports.models import *
from django.db.models import Q
from openpyxl import load_workbook
from base import views

def import_units_template(request):     
    if not request.user.is_authenticated:
        messages.error(request, "You are not authenticated to perform this action")
        return redirect("report-home-url")
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "Import Units Template"
        ws.append(["Unit Code", "Unit Name", "Teaching Hours Per Class", "Teaching Hours Per Week", "Teaching Hours Per Term"])
        ws["A1"].font = Font(bold=True)
        ws["B1"].font = Font(bold=True)
        ws["C1"].font = Font(bold=True)
        ws["D1"].font = Font(bold=True) 
        ws["E1"].font = Font(bold=True)
        
        response = HttpResponse(content_type="application/ms-excel")
        response["Content-Disposition"] = "attachment; filename=Import_units_template.xlsx"
        wb.save(response)
        return response
    
def import_courses_template(request):     
    if not request.user.is_authenticated:
        messages.error(request, "You are not authenticated to perform this action")
        return redirect("report-home-url")
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "Import Courses Template"
        ws.append(["Course Code", "Course Name"])
        ws["A1"].font = Font(bold=True)
        ws["B1"].font = Font(bold=True)
        
        response = HttpResponse(content_type="application/ms-excel")
        response["Content-Disposition"] = "attachment; filename=Import_courses_template.xlsx"
        wb.save(response)
        return response
    
def import_students_template(request):     
    if not request.user.is_authenticated:
        messages.error(request, "You are not authenticated to perform this action")
        return redirect("report-home-url")
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "Import Students Template"
        ws.append(
            [
                "Registration Number",
                "Full Name",
                "Course Code",
            ]
        )
        ws["A1"].font = Font(bold=True)
        ws["B1"].font = Font(bold=True)
        ws["C1"].font = Font(bold=True)
        
     
        
        response = HttpResponse(content_type="application/ms-excel")
        response["Content-Disposition"] = "attachment; filename=Import_students_template.xlsx"
        wb.save(response)
        return response

def generate_excel_for_student_reports(request):
    if not request.user.is_authenticated:
        messages.error(request, "You are not authenticated to perform this action")
        return redirect("report-home-url")
    else:
        series = Series.objects.get(is_active=True)
        wb = Workbook()
        ws = wb.active
        ws.title = "List of student reports"
        ws.append(
            [
                "Registration Number",
                "Student Name",
                "Course",
                "Series",
            ]
        )
        for student in StudentExam.objects.filter(is_active=True):
            ws.append(
                [
                    student.registration_number,
                    f"{student.name}",
                    f"{student.course.course_name} {student.course.course_code}",
                    series.name,
                ]
            )
        ws["A1"].font = Font(bold=True)
        ws["B1"].font = Font(bold=True)
        ws["C1"].font = Font(bold=True)
        ws["D1"].font = Font(bold=True)
        response = HttpResponse(content_type="application/ms-excel")
        response["Content-Disposition"] = (
            "attachment; filename=Active_student_reports.xlsx"
        )
        wb.save(response)
        return response


def generate_excel_for_course_reports(request):
    if not request.user.is_authenticated:
        messages.error(request, "You are not authenticated to perform this action")
        return redirect("report-home-url")
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "List of course reports"
        ws.append(["Course Name:", "Course Code:", "Units:"])
        for course in Course.objects.all():
            units = ", ".join(
                [unit.unit.unit_name for unit in course.courseunit_set.all()]
            )
            ws.append([course.course_name, course.course_code, units])
        ws["A1"].font = Font(bold=True)
        ws["B1"].font = Font(bold=True)
        ws["C1"].font = Font(bold=True)
        response = HttpResponse(content_type="application/ms-excel")
        response["Content-Disposition"] = "attachment; filename=Course_reports.xlsx"
        wb.save(response)
        return response


def generate_excel_for_unit_reports(request):
    if not request.user.is_authenticated:
        messages.error(request, "You are not authenticated to perform this action")
        return redirect("report-home-url")
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "List of unit reports"
        ws.append(["Unit Name", "Unit Code", "Courses"])
        for unit in Unit.objects.all():
            courses = ", ".join(
                [course.course.course_name for course in unit.courseunit_set.all()]
            )
            ws.append([unit.unit_name, unit.unit_code, courses])
        ws["A1"].font = Font(bold=True)
        ws["B1"].font = Font(bold=True)
        ws["C1"].font = Font(bold=True)
        response = HttpResponse(content_type="application/ms-excel")
        response["Content-Disposition"] = "attachment; filename=Unit_reports.xlsx"
        wb.save(response)
        return response

def generate_excel_clock_history_trainer(request, trainer_id):
    trainer = Trainer.objects.get(id=trainer_id)
    start_date = views.DatePicker.start_date
    end_date = views.DatePicker.end_date
    wb = Workbook()
    ws = wb.active
    ws.title = "Teaching Attendance History"

    ws.append([f"NAME: {trainer.name}"])
    ws.append([f"ID NUMBER: {trainer.id_number}"])
    ws.append(["UNIT NAME", "UNIT CODE", "CLASS DATE", "CLOCK IN TIME", "CLOCK OUT TIME", "ROLL", "DURATION"])
    clocks = TeachingAttendance.objects.filter(
            Q(clock_in__date__gte=start_date) &
            Q(clock_in__date__lte=end_date) &
            Q(trainer=trainer) &
            Q(is_clocked_out=True)
        )
    if clocks.exists():
            for clo in clocks:
                hours = clo.time_taken // 60
                minutes = clo.time_taken % 60
                time_taken = f"{hours} Hrs {minutes} Mins"
                row = [
                    clo.unit.unit_name,
                    clo.unit.unit_code,
                    clo.clock_in.date(),
                    str(clo.clock_in.astimezone().time()),
                    str(clo.clock_out.astimezone().time()),
                    clo.roll,
                    str(time_taken)
                ]
                ws.append(row)

            bold_font = Font(bold=True, condense=True)
            bold_font2 = Font(bold=True, condense=True)
            ws["A1"].font = bold_font
            ws["A2"].font = bold_font
            ws["A3"].font = bold_font2
            ws["B3"].font = bold_font2
            ws["C3"].font = bold_font2
            ws["D3"].font = bold_font2
            ws["E3"].font = bold_font2
            ws["F3"].font = bold_font2
            ws["G3"].font = bold_font2
            response = HttpResponse(
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = (f'attachment; filename="Teaching Attendance history for {trainer.id_number}'
                                               f'-attendance.xlsx"')
            with BytesIO() as stream:
                wb.save(stream)
                response.write(stream.getvalue())
            return response
    else:
            messages.error(request, 'No teaching attendance records found for this trainer.')
            return redirect('clock-history-url')


def generate_excel_clock_history_all(request):
    global minutes, hours
    try:
            start_date = views.DatePicker.start_date
            end_date = views.DatePicker.end_date
            wb = Workbook()
            ws = wb.active
            ws.title = "Teaching Attendance History"
            ws.append([f"TEACHING ATTENDANCE BETWEEN {start_date} AND {end_date}"])
            ws.append(
                ["TRAINER NAME", "ID NUMBER", "UNIT NAME", "UNIT CODE", "CLASS DATE", "CLOCK IN TIME", "CLOCK OUT TIME",
                 "ROLL", "DURATION"])
            clocks = TeachingAttendance.objects.filter(
                Q(clock_in__date__gte=start_date) &
                Q(clock_in__date__lte=end_date) &
                Q(is_clocked_out=True)
            )
            if clocks.exists():
                for clo in clocks:
                    if clo.time_taken:
                        hours = clo.time_taken // 60
                        minutes = clo.time_taken % 60
                    time_taken = f"{hours} Hrs {minutes} Mins"
                    row = [
                        clo.trainer.name,
                        clo.trainer.id_number,
                        clo.unit.unit_name,
                        clo.unit.unit_code,
                        clo.clock_out.astimezone().date(),
                        clo.clock_in.astimezone().time(),
                        clo.clock_out.astimezone().time() if clo.is_clocked_out else 'PENDING',
                        clo.roll,
                        str(time_taken)
                    ]
                    ws.append(row)

                bold_font = Font(bold=True, condense=True)
                bold_font2 = Font(bold=True, condense=True)
                ws["A1"].font = bold_font
                ws["A2"].font = bold_font2
                ws["B2"].font = bold_font2
                ws["C2"].font = bold_font2
                ws["D2"].font = bold_font2
                ws["E2"].font = bold_font2
                ws["F2"].font = bold_font2
                ws["G2"].font = bold_font2
                ws["H2"].font = bold_font2
                ws["I2"].font = bold_font2
                response = HttpResponse(
                    content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
                )
                response['Content-Disposition'] = (f'attachment; filename="Teaching Attendance Between {start_date} '
                                                   f'and {end_date}.xlsx"')
                with BytesIO() as stream:
                    wb.save(stream)
                    response.write(stream.getvalue())
                return response
    except TeachingAttendance.DoesNotExist:
            messages.error(request, 'Invalid request')
            return redirect('clock-history-all-url')
 
