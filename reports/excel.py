from django.http import HttpResponse
from django.shortcuts import redirect, render
from django.contrib import messages
from openpyxl import Workbook
from openpyxl.styles import Font
from io import BytesIO
from .models import *
from django.db.models import Q
import zipfile
from openpyxl import load_workbook


def generate_excel_for_student_reports(request):
    if not request.user.is_authenticated:
        messages.error(request, "You are not authenticated to perform this action")
        return redirect("report-home-url")
    else:
        series = Series.objects.get(is_active=True)
        wb = Workbook()
        ws = wb.active
        ws.title = "List of student reports"
        ws.append(
            [
                "Registration Number",
                "Student Name",
                "Course",
                "Series",
            ]
        )
        for student in StudentExam.objects.filter(is_active=True):
            ws.append(
                [
                    student.registration_number,
                    f"{student.name}",
                    f"{student.course.course_name} {student.course.course_code}",
                    series.name,
                ]
            )
        ws["A1"].font = Font(bold=True)
        ws["B1"].font = Font(bold=True)
        ws["C1"].font = Font(bold=True)
        ws["D1"].font = Font(bold=True)
        response = HttpResponse(content_type="application/ms-excel")
        response["Content-Disposition"] = (
            "attachment; filename=Active_student_reports.xlsx"
        )
        wb.save(response)
        return response


def generate_excel_for_course_reports(request):
    if not request.user.is_authenticated:
        messages.error(request, "You are not authenticated to perform this action")
        return redirect("report-home-url")
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "List of course reports"
        ws.append(["Course Name:", "Course Code:", "Units:"])
        for course in Course.objects.all():
            units = ", ".join(
                [unit.unit.unit_name for unit in course.courseunit_set.all()]
            )
            ws.append([course.course_name, course.course_code, units])
        ws["A1"].font = Font(bold=True)
        ws["B1"].font = Font(bold=True)
        ws["C1"].font = Font(bold=True)
        response = HttpResponse(content_type="application/ms-excel")
        response["Content-Disposition"] = "attachment; filename=Course_reports.xlsx"
        wb.save(response)
        return response


def generate_excel_for_unit_reports(request):
    if not request.user.is_authenticated:
        messages.error(request, "You are not authenticated to perform this action")
        return redirect("report-home-url")
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "List of unit reports"
        ws.append(["Unit Name", "Unit Code", "Courses"])
        for unit in Unit.objects.all():
            courses = ", ".join(
                [course.course.course_name for course in unit.courseunit_set.all()]
            )
            ws.append([unit.unit_name, unit.unit_code, courses])
        ws["A1"].font = Font(bold=True)
        ws["B1"].font = Font(bold=True)
        ws["C1"].font = Font(bold=True)
        response = HttpResponse(content_type="application/ms-excel")
        response["Content-Disposition"] = "attachment; filename=Unit_reports.xlsx"
        wb.save(response)
        return response


def generate_excel_for_marks_per_course_reports(request, series_id, course_id):
    if not request.user.is_authenticated:
        messages.error(request, "You are not authenticated to perform this action")
        return redirect("report-home-url")
    course = Course.objects.get(id=course_id)
    series = Series.objects.get(id=series_id)
    wb = Workbook()
    ws = wb.active
    ws.title = f"Marks per course"
    ws.append(["Course Name:", course.course_name])
    ws.append(["Series Name:", series.name])
    ws.append(["Student Name", "Registration Number", "Unit", "CAT Mark", "Exam Mark", "Project Mark"])
    for mark in Mark.objects.filter(
        student__course=course,
        series=series
    ):
        ws.append([
            f"{mark.student.name}",
            mark.student.registration_number,
            mark.unit.unit_name,
            mark.cat_mark,
            mark.exam_mark,
            mark.project_mark,
        ])
    for col in ws.iter_cols(min_row=1, max_row=3, min_col=1, max_col=6):
        for cell in col:
            cell.font = Font(bold=True)
    response = HttpResponse(content_type="application/ms-excel")
    response["Content-Disposition"] = "attachment; filename=Marks_per_course_reports.xlsx"
    wb.save(response)
    return response


def generate_excel_for_marks_per_student_reports(request, series_id, student_id):
    if not request.user.is_authenticated:
        messages.error(request, "You are not authenticated to perform this action")
        return redirect("report-home-url")
    series = Series.objects.get(id=series_id)
    try:
        student = StudentExam.objects.get(id=student_id)
        wb = Workbook()
        ws = wb.active
        ws.title = f"Marks for {student.name}"
        ws.append(["Student Name:", f"{student.name}"])
        ws.append(["Registration Number:", f"{student.registration_number}"])
        ws.append(["Series Name:", series.name])
        ws.append(["Unit", "CAT Mark", "Exam Mark", "Project Mark"])
        for mark in Mark.objects.filter(series=series, student=student):
            ws.append([mark.unit.unit_name, mark.cat_mark, mark.exam_mark, mark.project_mark])
        for col in ws.iter_cols(min_row=1, max_row=4, min_col=1, max_col=5):
            for cell in col:
                cell.font = Font(bold=True)
        response = HttpResponse(content_type="application/ms-excel")
        response["Content-Disposition"] = f"attachment; filename=Marks_for_{student.name}.xlsx"
        wb.save(response)
        return response
    except StudentExam.DoesNotExist:
        messages.error(request, f"Student not found!")
        return redirect('report-mark-reports-url')


def generate_excel_for_marks_per_unit_reports(request, series_id, unit_id):
    if not request.user.is_authenticated:
        messages.error(request, "You are not authenticated to perform this action")
        return redirect("report-home-url")
    unit = Unit.objects.get(id=unit_id)
    series = Series.objects.get(id=series_id)
    wb = Workbook()
    ws = wb.active
    ws.title = f"Marks for unit ({unit.unit_name})"
    ws.append(["Unit Name:", f"{unit.unit_name}({unit.unit_code})"])
    ws.append(["Series Name:", series.name])
    ws.append(["Student Name", "Registration Number", "CAT Mark", "Exam Mark", "Project Mark"])
    for mark in Mark.objects.filter(unit=unit, series=series):
        ws.append([
            f"{mark.student.name}",
            mark.student.registration_number,
            mark.cat_mark,
            mark.exam_mark,
            mark.project_mark,
        ])
    for col in ws.iter_cols(min_row=1, max_row=3, min_col=1, max_col=5):
        for cell in col:
            cell.font = Font(bold=True)
    response = HttpResponse(content_type="application/ms-excel")
    response["Content-Disposition"] = f"attachment; filename=Marks_for_unit_{unit.unit_name}.xlsx"
    wb.save(response)
    return response


def generate_excel_for_marks_per_series_reports(request, series_id):
    if not request.user.is_authenticated:
        messages.error(request, "You are not authenticated to perform this action")
        return redirect("report-home-url")
    series = Series.objects.get(id=series_id)
    wb = Workbook()
    ws = wb.active
    ws.title = f"Marks for series"
    ws.append(["Series Name:", series.name])
    ws.append(["Student Name", "Registration Number", "Unit", "CAT Mark", "Exam Mark", "Project Mark"])
    for mark in Mark.objects.filter(series=series):
        ws.append([
            f"{mark.student.name}",
            mark.student.registration_number,
            mark.unit.unit_name,
            mark.cat_mark,
            mark.exam_mark,
            mark.project_mark,
        ])
    for col in ws.iter_cols(min_row=1, max_row=2, min_col=1, max_col=6):
        for cell in col:
            cell.font = Font(bold=True)
    response = HttpResponse(content_type="application/ms-excel")
    response["Content-Disposition"] = f"attachment; filename=Marks_for_series_{series.name}.xlsx"
    wb.save(response)
    return response
    

def import_couses_template(request):     
    if not request.user.is_authenticated:
        messages.error(request, "You are not authenticated to perform this action")
        return redirect("report-home-url")
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "Import Courses Template"
        ws.append(["Course Code", "Course Name"])
        ws["A1"].font = Font(bold=True)
        ws["B1"].font = Font(bold=True)
        
        response = HttpResponse(content_type="application/ms-excel")
        response["Content-Disposition"] = "attachment; filename=Import_courses_template.xlsx"
        wb.save(response)
        return response
    
def import_units_template(request):     
    if not request.user.is_authenticated:
        messages.error(request, "You are not authenticated to perform this action")
        return redirect("report-home-url")
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "Import Units Template"
        ws.append(["Unit Code:", "Unit Name:"])
        ws["A1"].font = Font(bold=True)
        ws["B1"].font = Font(bold=True)
        
        response = HttpResponse(content_type="application/ms-excel")
        response["Content-Disposition"] = "attachment; filename=Import_units_template.xlsx"
        wb.save(response)
        return response

def import_students_template(request):     
    if not request.user.is_authenticated:
        messages.error(request, "You are not authenticated to perform this action")
        return redirect("report-home-url")
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "Import Students Template"
        ws.append(
            [
                "Registration Number",
                "Full Name",
                "Course Code",
            ]
        )
        ws["A1"].font = Font(bold=True)
        ws["B1"].font = Font(bold=True)
        ws["C1"].font = Font(bold=True)
        
     
        
        response = HttpResponse(content_type="application/ms-excel")
        response["Content-Disposition"] = "attachment; filename=Import_students_template.xlsx"
        wb.save(response)
        return response
