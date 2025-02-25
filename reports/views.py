from io import BytesIO
import json
from django.views.decorators.csrf import csrf_exempt
from django.contrib.auth.decorators import login_required
from django.shortcuts import render
from django.contrib import messages
from openpyxl import load_workbook
from reports.forms import *
from reports.models import *
from django.shortcuts import get_object_or_404, redirect
from django.urls import reverse
import csv
from django.http import HttpResponse, HttpResponseRedirect, JsonResponse
from django.forms import modelformset_factory
from .forms import MarksFormSet  # Import the formset


# Create your views here.
@login_required(login_url='signin-url')
def profile(request):
    form = UserForm(instance=request.user)
    return render(request, 'reports/profile.html', {'form_profile': form})

@login_required(login_url='signin-url')
def profile_edit(request):
    if request.method == 'POST':
            form = UserForm(request.POST, instance=request.user)
            if form.is_valid():
                form.save() 
                messages.success(request, 'You edited your profile successfully')
                return redirect('report-profile-url')
            else:
                messages.error(request, 'Please correct the errors')
    else:
            form = UserForm(instance=request.user)
    return render(request, 'reports/profile.html', {'form_profile': form})
     
@login_required(login_url='signin-url')
def report_home(request):
    series = None
    try:
        series = Series.objects.get(is_active=True)
    except Series.DoesNotExist:
        pass
    context = {"series": series}
    return render(request, "reports/home.html", context)

@login_required(login_url='signin-url')
def students_all(request):
    students = StudentExam.objects.filter(is_active=True)
    return render(request, "reports/students.html", {"students": students})

@login_required(login_url='signin-url')
def student_detail(request, student_id):
    student = get_object_or_404(StudentExam, id=student_id)
    return render(request, "reports/student_details.html", {"student": student})

@login_required(login_url='signin-url')
def add_student(request):
    if request.method == "POST":
        form = StudentForm(request.POST)
        if form.is_valid():
            form.save()             
            messages.error(request, "Student saved successfully!.")
            return HttpResponseRedirect("/report/student/mark/add/" + str(form.instance.id) + "/")
            # return redirect("report-students-all-url")
    else:
        form = StudentForm()
    return render(request, "reports/student_add.html", {"form": form})
@login_required(login_url='signin-url')
def mark_add_after_student_add(request, id):
    form = MarkFormAfterStudentAdd()
    student = get_object_or_404(StudentExam, id=id)
    marks = Mark.objects.filter(student=student).values_list('unit', flat=True)
    units = CourseUnit.objects.filter(course=student.course).exclude(unit__in=marks)
    if request.method == "POST":
        form = MarkFormAfterStudentAdd(request.POST)
        unit = request.POST.get("unit")
        unit = Unit.objects.get(id=unit)
        if form.is_valid():
            form.instance.student = student
            form.instance.series = Series.objects.get(is_active=True)
            form.instance.unit = unit
            form.save()
            messages.success(request, f"Marks for {unit.unit_name} saved successfully!.")
            return HttpResponseRedirect("/report/student/mark/add/" + str(id) + "/")
    return render(request, "reports/marks_add_after_student_reg.html", {"form": form, "student": student, "units": units})
    

@login_required(login_url='signin-url')
def edit_student(request, student_id):
    student = get_object_or_404(StudentExam, id=student_id)
    if request.method == "POST":
        form = StudentForm(request.POST, instance=student)
        if form.is_valid():
            form.save()
            return redirect("report-students-all-url")
    else:
        form = StudentForm(instance=student)
    return render(request, "reports/student_edit.html", {"form": form})

@login_required(login_url='signin-url')
def delete_student(request, student_id):
    student = get_object_or_404(StudentExam, id=student_id)
    if request.method == "DELETE":
        student.delete()
        return JsonResponse(
            {"success": True, "message": "Student deleted successfully"}
        )
    return JsonResponse({"success": False, "message": "Failed to delete student"})


@login_required(login_url='signin-url')
def mark_entry(request):
    return render(request, "reports/mark_entry.html")

@login_required(login_url='signin-url')
def search(request):
    if request.method == "POST":
        data = json.loads(request.body)
        query = data.get("query", "")
        students = StudentExam.objects.filter(registration_number__icontains=query)
        results = [
            {
                "id": student.id,
                "name": f"{student.name}",
                "regno": student.registration_number,
            }
            for student in students
        ]
        return JsonResponse({"success": True, "data": results})
    return JsonResponse({"success": False, "message": "Invalid request method"})

@login_required(login_url='signin-url')
def mark_add(request, id):
    student = get_object_or_404(StudentExam, id=id)
    try:
        series = Series.objects.get(is_active=True)
    except Series.DoesNotExist:
        messages.error(request, "No active Series, Please Create One or Extend the current Series End Dates")
        return redirect("report-series-url")
    units = CourseUnit.objects.filter(course=student.course)
    marks = Mark.objects.filter(student=student, series=series)
    # Check if marks already exist for this student
    if marks.exists():
        messages.error(request, f"Marks for {student.name} already exist. Please Edit Below")
        return HttpResponseRedirect("/report/mark/edit/" + str(id) + "/")
    elif not marks.exists():
        messages.error(request, "No marks entered for this student. Please Add Marks Below")
        return HttpResponseRedirect("/report/student/mark/add/" + str(id) + "/")
    context = {
        "student": student,
        "units": units,
        "series": series,
    }
    return render(request, "reports/marks_add.html", context)

@login_required(login_url='signin-url')
def marks_edit(request, id):
    student = get_object_or_404(StudentExam, id=id)
    units = CourseUnit.objects.filter(course=student.course)
    series = Series.objects.get(is_active=True)

    # Retrieve marks as a dictionary {unit.id: mark object}
    marks = {
        mark.unit.id: mark
        for mark in Mark.objects.filter(student=student, series=series)
    }

    if request.method == "POST":
        for unit in units:
            cat_mark = request.POST.get(f"cat_mark_{unit.unit.id}")
            exam_mark = request.POST.get(f"exam_mark_{unit.unit.id}")
            project_mark = request.POST.get(f"project_mark_{unit.unit.id}")

            if unit.unit.id in marks:
                # Update existing marks
                mark = marks[unit.unit.id]
                mark.cat_mark = cat_mark
                mark.exam_mark = exam_mark
                mark.project_mark = project_mark
                mark.save()
            else:
                # Create a new mark entry if not present
                Mark.objects.create(
                    student=student,
                    series=series,
                    unit=unit.unit,
                    cat_mark=cat_mark,
                    exam_mark=exam_mark,
                    project_mark=project_mark,
                )

        messages.success(request, "Marks updated successfully!")
        return redirect(
            "report-mark-entry-url"
        )  # Redirect to prevent duplicate form submission

    context = {
        "student": student,
        "units": units,
        "series": series,
        "marks": marks,  # Pass as a dictionary
    }
    return render(request, "reports/marks_edit.html", context)

@login_required(login_url='signin-url')
def marks_reports(request):
    series = Series.objects.all
    units = Unit.objects.all
    courses = Course.objects.all
    students = StudentExam.objects.all
    context = {"series": series, "units": units, "courses": courses, "students": students}
    return render(request, "reports/marks_report.html", context)

@login_required(login_url='signin-url')
def series_view(request):
    series = Series.objects.all()
    context = {"serieses": series}
    return render(request, "reports/series.html", context)

@login_required(login_url='signin-url')
def add_series(request):
    if request.method == "POST":
        form = SeriesAddForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect("report-series-url")
    else:
        form = SeriesAddForm()
    return render(request, "reports/series_add.html", {"form": form})

@login_required(login_url='signin-url')
def edit_series(request, series_id):
    series = get_object_or_404(Series, id=series_id)
    if request.method == "POST":
        form = SeriesAddForm(request.POST, instance=series)
        if form.is_valid():
            form.save()
            return redirect("report-series-url")
    else:
        form = SeriesAddForm(instance=series)
    return render(request, "reports/series_edit.html", {"form": form})

@login_required(login_url='signin-url')
def delete_series(request, series_id):
    series = get_object_or_404(Series, id=series_id, is_active=True)
    if request.method == "DELETE":
        series.delete()
        return JsonResponse({"success": True, "message": "Series deleted successfully"})
    return JsonResponse({"success": False, "message": "Failed to delete Series"})


@login_required(login_url='signin-url')
def import_students_from_excel(request):
    if not request.user.is_authenticated:
        messages.error(request, "You are not authenticated to perform this action")
        return redirect("report-home-url")

    if request.method == "POST" and request.FILES.get("file"):
        excel_file = request.FILES["file"]
        wb = load_workbook(filename=BytesIO(excel_file.read()))
        ws = wb.active

        for row in ws.iter_rows(min_row=2, values_only=True):
            row_values = list(row)  # Convert to list
            if len(row_values) < 3 or not all(row_values):  
                continue  # Skip rows with missing values
            
            registration_number, name, course_code = row_values[:3]  

            if not registration_number or not course_code:
                continue  # Skip invalid rows
            
            course = Course.objects.filter(course_code=course_code).first()
            if course and not StudentExam.objects.filter(registration_number=registration_number).exists():
                std = StudentExam.objects.create(
                    name=name,
                    registration_number=registration_number,
                    course=course
                )
                std.save()

        messages.success(request, "Students imported successfully")
        return redirect("report-students-all-url")

    return render(request, "reports/import_students.html")


